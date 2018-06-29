import xlwt
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy
from common_utils import arial10
from common_utils.str_utils import StrUtils


class SaveToFile:

    @staticmethod
    # extra_params format : Its a list of tuples [('param description', param value)...]
    def format_gridsearchcv_result(grid_result, extra_params=None):
        list_to_export = []

        # for key in grid_result.cv_results_['params'].items()]:

        ############ store the columns header
        # store the grid_results
        columns = ['score', 'std_deviation']
        descriptions = [*grid_result.cv_results_['params'][0]]  # https://www.python.org/dev/peps/pep-0448/
        if extra_params is not None:
            descriptions.extend([i[0] for i in extra_params])
        columns.extend(descriptions)
        list_to_export.append(columns)

        means = grid_result.cv_results_['mean_test_score']
        stds = grid_result.cv_results_['std_test_score']
        params = grid_result.cv_results_['params']
        extra_params_values = None
        if extra_params is not None:
            extra_params_values = [i[1] for i in extra_params]
        for current_mean, current_stdev, current_params in zip(means, stds, params):
            # store the columns header
            columns = [current_mean, current_stdev]
            # for key,val in [(key,val) for key, val in dct.items()
            columns.extend([val for key, val in current_params.items()])
            if extra_params_values is not None:
                columns.extend(extra_params_values)
            list_to_export.append(columns)

        return list_to_export

    # score, params

    def save_to_excel(filename, sheet, datalist, append=False, skip_first_line=False):
        if append:
            if filename.endswith('xlsx'):
                book = load_workbook(filename)
                sh = FitSheetWrapper(book[sheet])
            else:
                book = copy(open_workbook(filename))
                sh = FitSheetWrapper(book.get_sheet(sheet))
            nb_rows = len(sh._Worksheet__rows) - 1
        else:
            book = xlwt.Workbook(encoding='latin-1')
            sh = FitSheetWrapper(book.add_sheet(sheet))
            nb_rows = 0

        for current_line_nb, current_line in enumerate(datalist):
            # if it is not
            if skip_first_line and current_line_nb == 0:
                continue
            for current_column_nb, current_cell in enumerate(current_line):
                # print(current_line_nb, current_column_nb)
                # print(type(current_cell))
                sh.write(current_line_nb + nb_rows, current_column_nb, StrUtils.str_write(current_cell))

        book.save(filename)


class FitSheetWrapper(object):
    """Try to fit columns to max size of any entry.
    To use, wrap this around a worksheet returned from the
    workbook's add_sheet method, like follows:

        sheet = FitSheetWrapper(book.add_sheet(sheet_name))

    The worksheet interface remains the same: this is a drop-in wrapper
    for auto-sizing columns.
    """
    def __init__(self, sheet):
        self.sheet = sheet
        self.widths = dict()

    def write(self, r, c, label='', *args, **kwargs):
        self.sheet.write(r, c, label, *args, **kwargs)
        width = int(arial10.fitwidth(label))
        if width > self.widths.get(c, 0):
            self.widths[c] = width
            self.sheet.col(c).width = width

    def __getattr__(self, attr):
        return getattr(self.sheet, attr)